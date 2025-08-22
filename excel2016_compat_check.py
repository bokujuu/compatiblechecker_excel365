#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel 2016 互換チェック:
- 対象: .xlsx / .xlsm のセル数式およびブック定義名（Named Formula）
- 出力: 同一フォルダに `<元ファイル名>_2016_compat_report.md`
- 目的: 2016に存在しない関数の使用箇所を特定し、代替案を提示
"""

import sys
import os
import re
from datetime import datetime

try:
    from openpyxl import load_workbook
except Exception as e:
    print("openpyxl の読み込みに失敗しました:", e)
    sys.exit(1)

# --- 2016に存在しない（＝検出したら「エラー」）関数一覧 ---
# 代替案も併記（レポートに出力）
INCOMPATIBLE_2016 = {
    # 検索/参照
    "XLOOKUP": "INDEX+MATCH または VLOOKUP",
    "XMATCH": "MATCH",
    # 動的配列（スピル）
    "FILTER": "Advanced Filter または ヘルパー列+INDEX/SMALL",
    "UNIQUE": "「重複の削除」やピボットテーブル",
    "SORT": "データタブの並べ替え、ヘルパー列+INDEX",
    "SORTBY": "ヘルパー列+RANK/INDEX",
    "SEQUENCE": "ROW(INDIRECT(\"1:\"&n)) 等の連番生成",
    "RANDARRAY": "RAND/RANDBETWEEN を複製",
    # 変数/ユーザー定義
    "LET": "補助セル or 名前定義で代替",
    "LAMBDA": "VBA のユーザー定義関数(UDF)",
    "MAP": "補助列展開 or VBA",
    "REDUCE": "SUMPRODUCT/配列式またはVBA",
    "SCAN": "累積計算は相対参照のSUM等で代替 or VBA",
    "MAKEARRAY": "ROW/INDIRECT + INDEX 等の組合せ",
    # テキスト操作（新）
    "TEXTSPLIT": "MID/LEFT/RIGHT + FIND/SEARCH または「区切り位置指定」",
    "TEXTBEFORE": "LEFT + FIND/SEARCH",
    "TEXTAFTER": "MID + FIND/SEARCH",
    # 配列整形/結合
    "TOCOL": "INDEX + ROW/INDIRECT",
    "TOROW": "INDEX + COLUMN/INDIRECT",
    "VSTACK": "Power Queryで結合 or 手動結合",
    "HSTACK": "Power Queryで結合 or 手動結合",
    "TAKE": "INDEX で範囲切り出し",
    "DROP": "INDEX/OFFSET で除外",
    "CHOOSECOLS": "INDEX で列抽出",
    "CHOOSEROWS": "INDEX で行抽出",
    "WRAPROWS": "OFFSET/INDEX の組合せ",
    "WRAPCOLS": "OFFSET/INDEX の組合せ",
    "EXPAND": "IFERROR 等でパディング",
    # 画像
    "IMAGE": "挿入→画像（関数では不可）",
}

# --- 2016 で「注意（環境依存）」の関数一覧 ---
# 旧Excel 2016のビルドでは未実装だが、後年の更新や他エディションでは存在する場合がある
WARN_2016 = {
    "TEXTJOIN": "CONCATENATE または & 連結",
    "CONCAT": "CONCATENATE または & 連結",
    "IFS": "入れ子のIF",
    "SWITCH": "CHOOSE または 入れ子のIF",
    "MAXIFS": "MAX(IF(...)) の配列式（Ctrl+Shift+Enter）",
    "MINIFS": "MIN(IF(...)) の配列式（Ctrl+Shift+Enter）",
    # 必要に応じて追加:
    # "FORECAST.ETS": "FORECAST 等で代替（機能差あり）",
}

# --- 解析対象拡張子 ---
SUPPORTED_EXT = {".xlsx", ".xlsm"}  # .xlsb は参考対応（式抽出が限定的）

# --- ユーティリティ: 数式から関数名の抽出 ---
_re_quotes = re.compile(r'\"[^\"]*\"')  # ダブルクォート内を除外
_re_funcs = re.compile(r'([@_A-Z][@A-Z0-9\._]*)\s*(?=\()', re.IGNORECASE)

def normalize_func_name(raw: str) -> str:
    """
    _xlfn.XLOOKUP / @XLOOKUP などを XLOOKUP に正規化
    """
    s = raw.strip()
    # 先頭の @_xlfn. 等を除去
    s = s.lstrip('@')
    if '.' in s:
        s = s.split('.')[-1]
    return s.upper()

def extract_functions(formula: str):
    """
    文字列リテラルを除去した上で、関数呼び出しパターンを抽出
    """
    if not formula or not isinstance(formula, str):
        return []
    f = formula
    # 先頭 '=' を除去
    if f.startswith('='):
        f = f[1:]
    # 文字列リテラルを一時除去（誤検出防止）
    f_wo_str = _re_quotes.sub('', f)
    funcs = []
    for m in _re_funcs.finditer(f_wo_str):
        funcs.append(normalize_func_name(m.group(1)))
    return funcs

def shorten(s: str, maxlen: int = 180) -> str:
    s = s.replace('\n', ' ').replace('\r', ' ')
    return s if len(s) <= maxlen else s[:maxlen] + " ..."

def analyze_xlsx_xlsm(path: str):
    # 参照エラーも拾うため data_only=False。定義名や参照先を辿る必要はない
    wb = load_workbook(path, data_only=False, read_only=False)
    incompatible_hits = []  # [(sheet, addr, formula, [funcs])]
    warn_hits = []          # [(sheet, addr, formula, [funcs])]
    all_found_incompat = set()
    all_found_warn = set()

    # --- セル数式 ---
    for ws in wb.worksheets:
        # 最大使用範囲のみを走査
        for row in ws.iter_rows(min_row=1, min_col=1,
                                max_row=ws.max_row or 1, max_col=ws.max_column or 1,
                                values_only=False):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    funcs = extract_functions(v)
                    bad = sorted({fn for fn in funcs if fn in INCOMPATIBLE_2016})
                    warn = sorted({fn for fn in funcs if fn in WARN_2016})
                    if bad:
                        incompatible_hits.append((ws.title, cell.coordinate, v, bad))
                        all_found_incompat.update(bad)
                    if warn:
                        warn_hits.append((ws.title, cell.coordinate, v, warn))
                        all_found_warn.update(warn)

    # --- 定義名（Named Formula）内の LAMBDA 等も検査 ---
    try:
        for dn in wb.defined_names.definedName:
            if dn.attr_text and dn.attr_text.startswith("="):
                funcs = extract_functions(dn.attr_text)
                bad = sorted({fn for fn in funcs if fn in INCOMPATIBLE_2016})
                warn = sorted({fn for fn in funcs if fn in WARN_2016})
                # 定義名自体はセル座標を持たないので仮想行として出す
                where = f"[DefinedName] {dn.name}"
                if bad:
                    incompatible_hits.append((where, "-", dn.attr_text, bad))
                    all_found_incompat.update(bad)
                if warn:
                    warn_hits.append((where, "-", dn.attr_text, warn))
                    all_found_warn.update(warn)
    except Exception:
        pass  # 定義名無し等は無視

    return incompatible_hits, warn_hits, sorted(all_found_incompat), sorted(all_found_warn)

def write_report_md(src_path: str, incompatible_hits, warn_hits, found_incompat, found_warn):
    base = os.path.splitext(os.path.basename(src_path))[0]
    out_path = os.path.join(os.path.dirname(src_path), f"{base}_2016_compat_report.md")

    lines = []
    lines.append(f"# Excel 2016 互換性レポート")
    lines.append("")
    lines.append(f"- 対象ファイル: `{os.path.basename(src_path)}`")
    lines.append(f"- 実行時刻: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")

    # 概要
    lines.append("## 概要")
    lines.append(f"- エラー（2016に存在しない関数）検出数: **{len(incompatible_hits)}** 箇所")
    lines.append(f"- 注意（2016で環境依存の関数）検出数: **{len(warn_hits)}** 箇所")
    lines.append("")

    # エラー詳細
    lines.append("## エラー（2016に存在しない関数）")
    if not incompatible_hits:
        lines.append("- なし")
    else:
        for (sheet, addr, formula, bad) in incompatible_hits:
            lines.append(f"- **{', '.join(bad)}** @ `{sheet}` ! `{addr}`")
            lines.append(f"  - 数式: `{shorten(formula)}`")
        lines.append("")

    # 注意詳細
    lines.append("## 注意（2016で一部環境のみの関数）")
    if not warn_hits:
        lines.append("- なし")
    else:
        for (sheet, addr, formula, warn) in warn_hits:
            lines.append(f"- **{', '.join(warn)}** @ `{sheet}` ! `{addr}`")
            lines.append(f"  - 数式: `{shorten(formula)}`")
        lines.append("")

    # 代替案（関数ごと）
    if found_incompat or found_warn:
        lines.append("## 参考：代替案（関数別）")
        if found_incompat:
            lines.append("### 2016に存在しない関数")
            for fn in found_incompat:
                alt = INCOMPATIBLE_2016.get(fn, "")
                lines.append(f"- **{fn}** → {alt}")
        if found_warn:
            lines.append("")
            lines.append("### 2016で環境依存の関数（注意）")
            for fn in found_warn:
                alt = WARN_2016.get(fn, "")
                lines.append(f"- **{fn}** → {alt}")
    else:
        lines.append("## 参考：代替案（関数別）")
        lines.append("- 対象関数の検出はありませんでした。")

    # 制限事項
    lines.append("")
    lines.append("## 制限事項")
    lines.append("- 本ツールはセル数式と定義名を対象とします。条件付き書式やデータの入力規則の内部式は未検査です。")
    lines.append("- .xlsb の式抽出は未対応です（必要な場合は一度 .xlsx へ保存してから実行してください）。")
    lines.append("- 関数のローカライズ（言語別関数名）には非対応です。一般的な英語関数名で判定します。")

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return out_path

def _ensure_utf8_stdout():
    # Windows の cmd でも日本語が文字化けしにくいよう UTF-8 に固定
    try:
        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8")
        if hasattr(sys.stderr, "reconfigure"):
            sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

def main():
    _ensure_utf8_stdout()
    if len(sys.argv) < 2:
        print("使い方: python excel2016_compat_check.py <Excelファイル.xlsx|xlsm> [...複数可]")
        sys.exit(1)

    exit_code = 0
    for src in sys.argv[1:]:
        ext = os.path.splitext(src)[1].lower()
        if not os.path.isfile(src):
            print(f"ファイルが見つかりません: {src}")
            exit_code = 1
            continue

        if ext not in SUPPORTED_EXT:
            print(f"未対応拡張子: {src} （対応: .xlsx/.xlsm）")
            exit_code = 1
            continue

        try:
            incompatible_hits, warn_hits, found_incompat, found_warn = analyze_xlsx_xlsm(src)
            out_path = write_report_md(src, incompatible_hits, warn_hits, found_incompat, found_warn)
            # 結果概要を標準出力
            print(f"[OK] レポート出力: {out_path}")
            print(f"  エラー: {len(incompatible_hits)} 箇所, 注意: {len(warn_hits)} 箇所")
            if incompatible_hits:
                exit_code = 0  # 検出はしてもエラー終了にはしない
        except Exception as e:
            print(f"[NG] 解析中に例外: {src} -> {e}")
            exit_code = 1

    sys.exit(exit_code)

if __name__ == "__main__":
    main()
